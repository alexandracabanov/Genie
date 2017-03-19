#!/usr/local/bin/python3

#some modules we need
import openpyxl 
import wikipedia
import sys
import argparse

#this is in local fs here
import mouseminesearch

#Deal with command line args
parser = argparse.ArgumentParser(description='Why Hello!  Thank you so much for reading me. I\'m a Gene Seacher Genie (get it?!?) and I automate the annoying job of searching wikipedia for genes. I make erros sometimes, so make sure to QA the output. I more recently started searching mousemine.org too! Even with the issues, I can save you some time, if you have been tasked by your Post-doc to "get a feel" for whether Genes are generally well known.')
parser.add_argument('-i','--input', help='File Name of Gene List. Required format is one Gene per line.',required=True)
parser.add_argument('-o','--output',help='Specifcy the output file name. ".xlsx" will be appended to all names given', required=True)

#important note that the default search mods is specified here 
parser.add_argument('-w','--wikisearchmod',default='Gene',help='A wikepedia search modifier can be specified to give wikipedia more accurate results. The search modifier will be appended to the wiki search for each gene. The default search modifier is "Gene". This is perhaps the only useful search modifier for primary use case I wrote this for, but could be useful if I want to use this app to search other things other than Genes. (default: "Gene" appended to search string)', required=False)
parser.add_argument('-s','--species',default='Mus Musculus',help='Species constraint for mousemine.org search. To mess with this, you really need to know what mousemine.org is expecting. Should probably leave alone for now... (default: "Mus Musculus")', required=False)
args = parser.parse_args()

#workbooksetup
wb = openpyxl.Workbook()
ws1 = wb.active 
ws1.title='Genie'
#ws1 = wb.create_sheet("WikiPedia Search")

#set wb name to be written
#append xlsx to given output name incase people forget
#obviously we may have some .xlsx.xlsx files and that's okay
wbname=args.output+'.xlsx'

#start some counters to keep track iterations in the loops 
rowcnt=1 #starts a 1 since its used for column writing 
pgerrorcnt=0
notfoundcnt=0
disambigcnt=0
successcnt=0
mouseminesuccess=0
mouseminefail=0


#setup the workbork header row and column withds
ws1.cell(row=rowcnt, column=1, value="Gene")
ws1.cell(row=rowcnt, column=2, value="Wiki Summary")
ws1.cell(row=rowcnt, column=3, value="Mousemine Summary")
ws1.cell(row=rowcnt, column=4, value="WikiURL")
ws1.cell(row=rowcnt, column=5, value="Jax URL")
ws1.cell(row=rowcnt, column=6, value="MousemineURL")
ws1.column_dimensions['A'].width = 10
ws1.column_dimensions['B'].width = 60
ws1.column_dimensions['C'].width = 60
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 20 
ws1.column_dimensions['F'].width = 20


#quick loop to format the header row just set above 
columncnt=1
while columncnt <=6:
 #setup the header row of the excel file
 ws1.cell(row=rowcnt, column=columncnt).fill = openpyxl.styles.PatternFill('solid', 'A9A9A9')
 ws1.cell(row=rowcnt, column=columncnt).font = openpyxl.styles.Font(bold=True)
 columncnt+=1 

#This section opens the file specified and counts all lines. 
#It then prints a message telling how many genes were loaded from the file specified
with open(args.input) as fcnt: 
 rowcnt=0 #reset rowcnt to be 0 so we can count total lines real quick 
 for linecnt in fcnt: 
   rowcnt+=1
 searchcnt=rowcnt #store rowcnt in searchcnt for later use
 print ("\nLoaded",rowcnt,"genes to query from file",args.input,"\n")

#This is the main loop of the 
rowcnt=1 #put it back to 1 so we can key off of it for writing cells and don't overwrite header
with open(args.input) as f: 
 for line in f: 
   rowcnt+=1 #for first loop, start at line 2

   #construct the query for wikipedia as variable "line"
   gene=line #pull out gene into its own variable since we're going to change line with a search modifier 
   line=args.wikisearchmod+' '+line
   
   #add gene to first column of row and format some rows
   ws1.cell(row=rowcnt, column=1, value=gene.upper()) 
   ws1.cell(row=rowcnt, column=1).font = openpyxl.styles.Font(bold=True)
   ws1.cell(row=rowcnt, column=1).alignment = openpyxl.styles.Alignment(vertical="top",wrap_text=True,shrink_to_fit=False)
   ws1.cell(row=rowcnt, column=3).alignment = openpyxl.styles.Alignment(vertical="top",wrap_text=True,shrink_to_fit=False)
   ws1.cell(row=rowcnt, column=1).number_format = '@' #make text format so excel doesn't make gene names dates -  http://openpyxl.readthedocs.io/en/default/_modules/openpyxl/styles/numbers.html
   
   #tell user what is being searched 
   print('Searching for \"',gene.strip(),'\" at mousemine.org on \"',args.species,'\" species. Result: ',sep='',end='')
   #perform the search on mousemine
   mouseminereturn=mouseminesearch.genefind(gene.strip(),args.species)  
   if mouseminereturn != 'Nothing':
    if args.species == 'Mus Musculus':
     jaxurl=mouseminesearch.jaxify(mouseminereturn["primaryIdentifier"])
    else: 
     jaxurl='Only Generated for Mus Musculus' 
    mousemineurl=mouseminesearch.mouselink(mouseminereturn["id"])
    mouseminedesc=mouseminereturn["description"]
    print('Found \"',mousemineurl,'\"',sep='')
    mouseminesuccess+=1
   else:
    jaxurl="No Search Result"
    mousemineurl="No Search Result"
    mouseminedesc="No Search Result"
    print('Not Found ',sep='')
    mouseminefail+=1
    
   if mouseminedesc is None:
    mouseminedesc='No Mousemine Description Present' 
   
   ws1.cell(row=rowcnt, column=3, value=mouseminedesc)
   ws1.cell(row=rowcnt, column=3).alignment = openpyxl.styles.Alignment(vertical="top",wrap_text=True,shrink_to_fit=False)
   ws1.cell(row=rowcnt, column=5).hyperlink = jaxurl
   ws1.cell(row=rowcnt, column=5).font = openpyxl.styles.Font(bold=True,underline='single',color='0645AD')
   ws1.cell(row=rowcnt, column=5).alignment = openpyxl.styles.Alignment(vertical="top",wrap_text=True,shrink_to_fit=False)
   ws1.cell(row=rowcnt, column=6).hyperlink = mousemineurl
   ws1.cell(row=rowcnt, column=6).font = openpyxl.styles.Font(bold=True,underline='single',color='0645AD')
   ws1.cell(row=rowcnt, column=6).alignment = openpyxl.styles.Alignment(vertical="top",wrap_text=True,shrink_to_fit=False)
   
   print('Searching \"',line.strip(),'\" on WikiPedia. Result: ',sep='',end='')
   #execute wikiepedia search withing try statement due to common errors we can expect
   try: 
     #store wikipedia search result in object searchresult
     searchresult=wikipedia.search(line.upper())
     #test if something was found 
     if len(searchresult) > 0:
       pagename=searchresult[0]
       summary=wikipedia.summary(pagename)
       page=wikipedia.page(pagename)
       url=page.url
       print('Found \"',pagename,'\"',sep='')
       ws1.cell(row=rowcnt, column=2, value=summary)
       ws1.cell(row=rowcnt, column=2).alignment = openpyxl.styles.Alignment(vertical="top",wrap_text=True,shrink_to_fit=False) #http://openpyxl.readthedocs.io/en/default/_modules/openpyxl/styles/alignment.html
       ws1.cell(row=rowcnt, column=4).hyperlink = url
       ws1.cell(row=rowcnt, column=4).font = openpyxl.styles.Font(bold=True,underline='single',color='0645AD')
       ws1.cell(row=rowcnt, column=4).alignment = openpyxl.styles.Alignment(vertical="top",wrap_text=True,shrink_to_fit=False)
       successcnt+=1
     else: 
      url="N/A"
      print('No Result')
      #ws1.cell(row=rowcnt, column=1, value=line.upper())
      ws1.cell(row=rowcnt, column=2, value='no search result on wikipedia')
      ws1.cell(row=rowcnt, column=4, value=url)
      notfoundcnt+=1
   #catch typical errors from wikipedia not finding result of querry
   except wikipedia.exceptions.PageError:
     url="N/A"
     print('Page Error')
     #ws1.cell(row=rowcnt, column=1, value=line.upper())
     ws1.cell(row=rowcnt, column=2, value='no search result on wikipedia')
     ws1.cell(row=rowcnt, column=4, value=url)
     pgerrorcnt+=1
   #catch similar error when search is deemeed too ambiguous by wikipedia 
   except wikipedia.exceptions.DisambiguationError: 
     url="N/A"
     print('Ambiguous Reference')
     #ws1.cell(row=rowcnt, column=1, value=line.upper())
     ws1.cell(row=rowcnt, column=2, value='ambiguous ref.')
     ws1.cell(row=rowcnt, column=4, value=url)
     disambigcnt+=1


print("")
print("###########Results############")
print('Searched Genes:',searchcnt)
print('MouseMine Results Found:', mouseminesuccess)
print('Wiki Results Found:', successcnt)
print('Wiki Not Found:', notfoundcnt)
print('Mousemine Not Found:', mouseminefail)
print('Wiki Errors:', disambigcnt+pgerrorcnt)

print("")
print('Saving results to \"',wbname,'\"',sep='') 
print('Do not forget to check it out!!!') 

wb.save(wbname)
